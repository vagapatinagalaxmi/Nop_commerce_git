import openpyxl

def numRows(file_name,sheet_name):
    Excel_File = openpyxl.load_workbook(file_name)#loading file
    sheet = Excel_File[sheet_name]#loding  sheet name
    return sheet.max_row # in the specific we getting max rows

def readData(file_name,sheet_name,row_num,col_num):
    Excel_File =openpyxl.load_workbook(file_name)#loading file
    sheet =Excel_File[sheet_name]# loding sheet name
    return sheet.cell(row = row_num,column=col_num).value# returing the values of given rows and column

def writeData(file_name,sheet_name,row_num,col_num, data):
    Excel_File =openpyxl.load_workbook(file_name)
    sheet=Excel_File[sheet_name]
    sheet.cell(row=row_num,column=col_num).value=data
    Excel_File.save(file_name)


#readData(lpgindata.xlsx,Sheet1,2,3)
